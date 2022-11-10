from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import random
import os
from django.core.files.base import ContentFile
from .models import DocumentUpload
from .models import FormData
from django.db.models.query_utils import Q
import csv


def create_form_history():
    """ Generate csv file """

    output_file = open("./assests/output.csv", "w")
    output_file_writer = csv.writer(output_file)

    datas = FormData.objects.all()

    output_file_writer.writerow(["s/n", "FIRST NAME", "LAST NAME", "AGE",
                                "EMAIL", "PHONE NUMBER", "STATE", "L.G.A", "CATEGORY", "SHOP NAME"])

    for idx, data in enumerate(datas, start=1):
        output_file_writer.writerow([idx,  data.first_name,  data.last_name, data.age,
                                    data.email, data.phone_number, data.state, data.lga, data.category, data.shop_name])

    output_file.close()

    unique = random.randint(0, 10000000)
    uploadedFile = DocumentUpload.objects.create(name="payment_register")

    f = open(f'./assests/output.csv', 'rb')
    uploadedFile.document.save(f"output-{unique}.csv", ContentFile(f.read()))
    url = uploadedFile.document.url

    #   Close & Remove the local generated file
    f.close()
    # os.remove(f'./assests/output.csv')

    return {"downloadUrl": url,  "filename": 'form_records'}


# def create_form_history() -> dict:
#     ''' Generate Excel report  '''

#     #   WRITE TO WORKBOOK
#     wb = load_workbook(filename='./assests/epay_default.XLSX')
#     redColor = Font(color="00FF0000")
#     greenColor = Font(color="00008000")
#     sheet = wb["Sheet1"]

#     sheet.cell(row=2, column=2).value = paymentname[0:30]
#     sheet.cell(row=3, column=2).value = "₦ "+str(paymentprice)
#     sheet.cell(row=4, column=2).value = totalWhoPaid
#     sheet.cell(row=5, column=2).value = totalWhoNotPaid

#     count = 8

#     for n, resident in enumerate(residents, start=1):
#         sheet.cell(row=count, column=1).value = n
#         sheet.cell(row=count, column=2).value = resident.estate_id
#         sheet.cell(
#             row=count, column=4).value = resident.get_full_name()
#         if resident.id in residentsWhoPaid:
#             sheet.cell(row=count, column=6).value = "PAID"
#             sheet.cell(row=count, column=6).font = greenColor
#         else:
#             sheet.cell(row=count, column=6).value = "NOT PAID"
#             sheet.cell(row=count, column=6).font = redColor

#         count += 1

#     unique = random.randint(0, 10000000)

#     wb.save(f'./assests/epay_{paymentname}_{unique}_{estate}.XLSX')

#     uploadedFile = DocumentUpload.objects.create(
#         name="payment_register",
#     )

#     f = open(f'./assests/epay_{paymentname}_{unique}_{estate}.XLSX', 'rb')
#     uploadedFile.document.save(
#         f"epay_{paymentname}_{unique}_{estate}.XLSX'", ContentFile(f.read()))

#     url = uploadedFile.document.url

#     #   Close & Remove the local generated file
#     f.close()
#     os.remove(f'./assests/epay_{paymentname}_{unique}_{estate}.XLSX')

#     return {
#         "downloadUrl": url,
#         "filename": 'form_records'
#     }

    # return {
    #     "downloadUrl": "https://images.pexels.com/photos/35537/child-children-girl-happy.jpg?cs=srgb&dl=pexels-bess-hamiti-35537.jpg&fm=jpg&w=4713&h=3168&_gl=1*fuheh*_ga*MjEwNTgxOTY3NC4xNjY3ODYwMzAy*_ga_8JE65Q40S6*MTY2Nzg2MDMxNC4xLjAuMTY2Nzg2MDMxNC4wLjAuMA..",
    #     "filename": 'Estate_payment_records'
    # }
